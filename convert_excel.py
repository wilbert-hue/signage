"""
Convert U.S. Digital Signage PIVOT Excel to value.json and volume.json.

Reads hierarchical rows from the Value and Volume sheets (Excel indent 0–3).
Parent rows that have leaf children include year totals at the same object level as
children so json-processor can treat them as aggregated records.
"""
import json

import openpyxl

EXCEL_FILE = "Copy of PIVOT - U.S. Digital Signage Market.xlsx"
VALUE_SHEET = "Value"
VOLUME_SHEET = "Volume"
HEADER_ROW = 18
DATA_START_ROW = 19
OUT_VALUE = "public/data/value.json"
OUT_VOLUME = "public/data/volume.json"


def parse_year_columns(ws, header_row: int) -> list[str]:
    """Return year keys like '2021' in column order until CAGR or blank."""
    years: list[str] = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            break
        if isinstance(v, str):
            s = v.strip()
            if s.upper() == "CAGR":
                break
            if s.isdigit():
                years.append(str(int(s)))
            else:
                break
        elif isinstance(v, (int, float)) and 2000 < float(v) < 2100:
            years.append(str(int(v)))
        else:
            break
    return years


def cell_number(val, decimals: int) -> float:
    if val is None:
        return 0.0
    try:
        x = float(val)
    except (TypeError, ValueError):
        return 0.0
    return float(round(x, decimals))


def read_data_rows(ws, years: list[str], data_start: int, value_decimals: int):
    rows = []
    for row_idx in range(data_start, ws.max_row + 1):
        label = ws.cell(row_idx, 1).value
        if label is None:
            continue

        cell = ws.cell(row_idx, 1)
        indent = int(cell.alignment.indent) if cell.alignment and cell.alignment.indent else 0

        year_data = {}
        has_data = False
        for i, yk in enumerate(years):
            val = ws.cell(row_idx, 2 + i).value
            if val is not None:
                year_data[yk] = cell_number(val, value_decimals)
                has_data = True
            else:
                year_data[yk] = cell_number(None, value_decimals)

        label = label.strip() if isinstance(label, str) else str(label)

        rows.append(
            {
                "row_idx": row_idx,
                "label": label,
                "indent": indent,
                "year_data": year_data if has_data else None,
            }
        )

    return rows


def build_json_from_rows(rows):
    result = {}

    scan_geo = None
    scan_seg = None
    sub_seg_has_children = set()
    for i, row in enumerate(rows):
        indent = row["indent"]
        if indent == 0:
            scan_geo = row["label"]
            scan_seg = None
        elif indent == 1:
            scan_seg = row["label"]
        elif indent == 2 and row["year_data"]:
            for j in range(i + 1, len(rows)):
                if rows[j]["indent"] == 3:
                    sub_seg_has_children.add((scan_geo, scan_seg, row["label"]))
                    break
                if rows[j]["indent"] <= 2:
                    break

    current_geo = None
    current_seg_type = None
    current_sub_seg = None

    for row in rows:
        label = row["label"]
        indent = row["indent"]
        year_data = row["year_data"]

        if indent == 0:
            current_geo = label
            current_seg_type = None
            current_sub_seg = None
            if current_geo not in result:
                result[current_geo] = {}
            continue

        if current_geo is None:
            continue

        if indent == 1:
            current_seg_type = label
            current_sub_seg = None
            if current_seg_type not in result[current_geo]:
                result[current_geo][current_seg_type] = {}
            continue

        if current_seg_type is None:
            continue

        geo_data = result[current_geo][current_seg_type]

        if indent == 2:
            current_sub_seg = label

            if year_data is None:
                continue

            has_children = (current_geo, current_seg_type, label) in sub_seg_has_children

            if has_children:
                if label not in geo_data:
                    geo_data[label] = {}
                for year_key, year_val in year_data.items():
                    geo_data[label][year_key] = year_val
            else:
                geo_data[label] = year_data

        elif indent == 3:
            if current_sub_seg is None or year_data is None:
                continue
            if current_sub_seg not in geo_data:
                geo_data[current_sub_seg] = {}
            geo_data[current_sub_seg][label] = year_data

    return result


def hardware_volume_ratios_by_year(value_json: dict, volume_json: dict, years: list[str]) -> dict[str, float]:
    """
    For each year, sum(volume leaves) / sum(value leaves) under Hardware — used only
    where the Volume sheet has no rows. Ratios drift by year because value vs volume
    growth differs in the source file.
    """
    kk = value_json.get("U.S.", {}).get("By Component", {}).get("Hardware", {})
    hv = volume_json.get("U.S.", {}).get("By Component", {}).get("Hardware", {})
    children = [c for c in kk if not str(c).isdigit()]
    out: dict[str, float] = {}
    if not children or not isinstance(hv, dict):
        return {y: 0.0 for y in years}
    for y in years:
        try:
            sv = sum(float(kk[c][y]) for c in children if isinstance(kk.get(c), dict))
            so = sum(float(hv[c][y]) for c in children if isinstance(hv.get(c), dict))
        except (KeyError, TypeError, ValueError):
            out[y] = 0.0
            continue
        out[y] = (so / sv) if sv > 0 else 0.0
    return out


def merge_volume_from_value(value_node, vol_node, ratios_by_year: dict[str, float]) -> dict:
    """
    Align volume tree with value tree. Prefer numbers from vol_node when present;
    otherwise scale value leaves by ratios_by_year[y].
    """
    if not isinstance(value_node, dict):
        return {}
    vol_node = vol_node if isinstance(vol_node, dict) else {}

    year_keys = sorted((k for k in value_node if str(k).isdigit()), key=int)
    skip = frozenset({"CAGR", "_aggregated", "_level"})
    child_keys = [k for k in value_node if k not in year_keys and k not in skip]

    if year_keys and not child_keys:
        out = {}
        for y in year_keys:
            if y in vol_node and isinstance(vol_node[y], (int, float)):
                out[y] = cell_number(vol_node[y], 3)
            else:
                r = ratios_by_year.get(y, 0.0)
                out[y] = cell_number(float(value_node[y]) * r, 3)
        return out

    out = {}
    for y in year_keys:
        if y in vol_node and isinstance(vol_node[y], (int, float)):
            out[y] = cell_number(vol_node[y], 3)
        elif isinstance(value_node.get(y), (int, float)):
            r = ratios_by_year.get(y, 0.0)
            out[y] = cell_number(float(value_node[y]) * r, 3)
    for ck in child_keys:
        out[ck] = merge_volume_from_value(value_node[ck], vol_node.get(ck), ratios_by_year)
    return out


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    value_ws = wb[VALUE_SHEET]
    years = parse_year_columns(value_ws, HEADER_ROW)
    if not years:
        raise SystemExit(f"No year columns found in row {HEADER_ROW} of {VALUE_SHEET}")

    print(f"Years: {years[0]}..{years[-1]} ({len(years)} columns)")

    value_rows = read_data_rows(value_ws, years, DATA_START_ROW, value_decimals=2)
    print(f"{VALUE_SHEET}: {len(value_rows)} data rows")

    value_json = build_json_from_rows(value_rows)

    volume_ws = wb[VOLUME_SHEET]
    vol_years = parse_year_columns(volume_ws, HEADER_ROW)
    if vol_years != years:
        print(f"Warning: {VOLUME_SHEET} years {vol_years} != {VALUE_SHEET} years {years}")

    volume_rows = read_data_rows(volume_ws, vol_years or years, DATA_START_ROW, value_decimals=3)
    print(f"{VOLUME_SHEET}: {len(volume_rows)} data rows")

    partial_volume = build_json_from_rows(volume_rows)
    wb.close()

    blend_by_year = hardware_volume_ratios_by_year(value_json, partial_volume, years)
    mid = years[len(years) // 2]
    print(
        f"Volume/value ratio fallback (Hardware children, {years[0]} / {years[-1]}): "
        f"{blend_by_year[years[0]]:.4f} / {blend_by_year[years[-1]]:.4f} (mid {mid}: {blend_by_year[mid]:.4f})"
    )

    volume_complete: dict = {}
    for geo, seg_types in value_json.items():
        pv_geo = partial_volume.get(geo, {})
        volume_complete[geo] = {}
        for seg_type, subtree in seg_types.items():
            if seg_type == "By Component":
                # Volume source only defines hardware units; omit Software/Services from volume entirely
                v_hw = subtree.get("Hardware")
                pv_hw = pv_geo.get(seg_type, {}).get("Hardware") if isinstance(pv_geo.get(seg_type), dict) else {}
                hw_subtree = v_hw if isinstance(v_hw, dict) else {}
                pv_hw_clean = pv_hw if isinstance(pv_hw, dict) else {}
                merged_hw = merge_volume_from_value(hw_subtree, pv_hw_clean, blend_by_year)
                volume_complete[geo][seg_type] = {"Hardware": merged_hw} if merged_hw else {}
            else:
                volume_complete[geo][seg_type] = merge_volume_from_value(
                    subtree, pv_geo.get(seg_type, {}), blend_by_year
                )

    for geo in value_json:
        print(f"  {geo}: {list(value_json[geo].keys())}")

    with open(OUT_VALUE, "w", encoding="utf-8") as f:
        json.dump(value_json, f, indent=2, ensure_ascii=False)
    print(f"Wrote {OUT_VALUE}")

    with open(OUT_VOLUME, "w", encoding="utf-8") as f:
        json.dump(volume_complete, f, indent=2, ensure_ascii=False)
    print(f"Wrote {OUT_VOLUME}")

    def structure_without_years(obj: dict) -> dict:
        meta = frozenset({"CAGR", "_aggregated", "_level"})
        if not isinstance(obj, dict):
            return {}
        child_keys = [
            k
            for k in obj
            if k not in meta and not (isinstance(k, str) and k.isdigit())
        ]
        return {k: structure_without_years(obj[k]) for k in child_keys if isinstance(obj.get(k), dict)}

    seg_payload = {g: structure_without_years(value_json[g]) for g in value_json}
    seg_path = "public/data/segmentation_analysis.json"
    with open(seg_path, "w", encoding="utf-8") as f:
        json.dump(seg_payload, f, indent=2, ensure_ascii=False)
    print(f"Wrote {seg_path}")

    # Quick verification: By Component > Hardware total vs children sum 2021
    bc = value_json.get("U.S.", {}).get("By Component", {})
    hw = bc.get("Hardware", {})
    children = [k for k in hw if not k.isdigit()]
    if children:
        s = sum(hw[c].get("2021", 0) for c in children if isinstance(hw.get(c), dict))
        print(f"\nCheck U.S. > By Component > Hardware 2021: total={hw.get('2021')} children_sum={round(s, 2)}")


if __name__ == "__main__":
    main()
